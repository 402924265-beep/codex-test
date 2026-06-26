from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from .config import TMP_DIR


MONTH_TOKENS = {
    1: ("jan", "january"),
    2: ("feb", "february"),
    3: ("mar", "march"),
    4: ("apr", "april"),
    5: ("may",),
    6: ("jun", "june"),
    7: ("jul", "july"),
    8: ("aug", "august"),
    9: ("sep", "sept", "september"),
    10: ("oct", "october"),
    11: ("nov", "november"),
    12: ("dec", "december"),
}


SAP_SUMMARY_LABELS = {
    "CS_DEPRECIATION": "折旧（含FC）",
    "CS_DIRECT LABOUR": "直接人工",
    "CS_DIRECT LABOR": "直接人工",
    "CS_SCRAP_VARIABLE": "可回收废料",
    "CS_FIX COST": "运营费",
    "CS_VARIABLE COST": "生产耗用品",
    "CS_OBSOLESCENCE": "存货跌价准备",
    "CS_RESELLING": "废品回收",
    "CS_FIX UTILITIES": "固定能源费",
    "CS_VARIABLE UTILITIES": "变动能源费",
    "CS_INDIRECT LABOUR": "间接人工成本-辅助人员",
    "CS_INDIRECT LABOR": "间接人工成本-辅助人员",
    "CS_FIXED LABOUR": "固定人工-白领",
    "CS_FIXED LABOR": "固定人工-白领",
    "CS_SEMIFIX": "半固定-班车/工作服",
    "CS_OVH_ADM": "分摊费用",
}


@dataclass
class SourceAccount:
    code: str
    desc_en: str = ""
    desc_cn: str = ""
    group: str = ""
    amount: float | None = None
    unit: float | None = None
    volume: float | None = None
    source: str = ""
    rows: list[int] = field(default_factory=list)
    duplicate_count: int = 1


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\xa0", " ").strip()


def number(value: Any) -> float | None:
    if value in (None, "", "#DIV/0!"):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return None


def account_code(value: Any) -> str:
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric.is_integer():
            raw_number = str(int(numeric))
            if re.fullmatch(r"\d{6}|\d{10}", raw_number):
                return raw_number
        return ""
    raw = text(value)
    match = re.search(r"(?<![\w.])(?:\d{10}|\d{6})(?![\w.])", raw)
    if match:
        return match.group(0)
    upper = raw.upper().strip()
    if upper.startswith("FIX"):
        return "FIX"
    if upper.startswith("VAR"):
        return "VAR"
    return ""


def safe_load_workbook(path: Path, *, read_only: bool = True, data_only: bool = True):
    try:
        return load_workbook(path, read_only=read_only, data_only=data_only)
    except ValueError as exc:
        message = str(exc)
        if "defined names" not in message and "assign names" not in message and "print titles" not in message and "#N/A" not in message:
            raise
        cleaned = clean_invalid_defined_names(path)
        return load_workbook(cleaned, read_only=read_only, data_only=data_only)


def clean_invalid_defined_names(path: Path) -> Path:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(f"{path}|{path.stat().st_mtime_ns}".encode("utf-8", errors="ignore")).hexdigest()[:12]
    output = TMP_DIR / f"{path.stem}.cleaned.{digest}.xlsx"
    if output.exists():
        return output

    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                data = _remove_bad_defined_names(data)
            zout.writestr(info, data)
    return output


def _remove_bad_defined_names(xml_bytes: bytes) -> bytes:
    ET.register_namespace("", "http://schemas.openxmlformats.org/spreadsheetml/2006/main")
    root = ET.fromstring(xml_bytes)
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    defined_names = root.find("main:definedNames", ns)
    if defined_names is None:
        return xml_bytes
    for node in list(defined_names):
        value = text(node.text)
        name = text(node.attrib.get("name"))
        if "#N/A" in value or value.upper() == "N/A" or (name in {"_xlnm.Print_Titles", "_xlnm.Print_Area"} and not value):
            defined_names.remove(node)
    if len(list(defined_names)) == 0:
        root.remove(defined_names)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def find_sheet(workbook, candidates: list[str]) -> str:
    exact = {name.strip().lower(): name for name in workbook.sheetnames}
    for candidate in candidates:
        found = exact.get(candidate.strip().lower())
        if found:
            return found
    for sheet in workbook.sheetnames:
        normalized = sheet.strip().lower()
        if any(candidate.strip().lower() in normalized for candidate in candidates):
            return sheet
    raise ValueError(f"找不到目标sheet，候选：{', '.join(candidates)}；实际sheet：{', '.join(workbook.sheetnames)}")


def detect_month_columns(ws, month: int) -> tuple[int, int, int]:
    tokens = MONTH_TOKENS[month]
    for row in range(1, min(ws.max_row, 20) + 1):
        value_col = None
        cpu_col = None
        for col in range(1, min(ws.max_column, 80) + 1):
            label = text(ws.cell(row, col).value).lower()
            if not label:
                continue
            has_month = any(token in label for token in tokens)
            if not has_month:
                continue
            if "cpu" in label:
                cpu_col = col
            elif "k€" in label or "k eur" in label or "keur" in label or "k " in label or label.endswith("k"):
                value_col = col
        if value_col:
            return row, value_col, cpu_col or value_col + 1
    raise ValueError(f"{ws.title} 中识别不到 {month} 月 K€ / CPU 列")


def detect_volume(ws, value_col: int) -> float | None:
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(max(1, value_col - 2), min(ws.max_column, value_col + 3) + 1):
            value = number(ws.cell(row, col).value)
            if value is not None and value > 100:
                return value
    return None


def extract_renta_accounts(path: Path, sheet_candidates: list[str], month: int) -> tuple[dict[str, SourceAccount], dict[str, Any]]:
    wb = safe_load_workbook(path, read_only=True, data_only=True)
    sheet_name = find_sheet(wb, sheet_candidates)
    ws = wb[sheet_name]
    if sheet_name.strip().upper() == "SAP_ACT_EUR":
        accounts, meta = extract_sap_act_eur_accounts(wb, path, sheet_name, month)
        wb.close()
        return accounts, meta
    header_row, value_col, cpu_col = detect_month_columns(ws, month)
    volume = detect_volume(ws, value_col)
    accounts: dict[str, SourceAccount] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(row)
        code_col_idx = _find_code_index(values)
        is_fc = _is_functional_currency_row(values)
        if code_col_idx is None and not is_fc:
            continue
        code = "__FC__" if is_fc else account_code(values[code_col_idx])
        amount = number(values[value_col - 1] if value_col - 1 < len(values) else None)
        unit = number(values[cpu_col - 1] if cpu_col - 1 < len(values) else None)
        if is_fc and amount is None and unit is None:
            continue
        desc_en = "Functional Currency Impact" if is_fc else text(values[code_col_idx + 1] if code_col_idx + 1 < len(values) else "")
        desc_cn = "汇率影响" if is_fc else text(values[code_col_idx + 2] if code_col_idx + 2 < len(values) else "")
        group = "Functional Currency Impact" if is_fc else text(values[4] if len(values) > 4 else "")
        if code not in accounts:
            accounts[code] = SourceAccount(
                code=code,
                desc_en=desc_en,
                desc_cn=desc_cn,
                group=group,
                amount=amount,
                unit=unit,
                volume=volume,
                source=f"{path.name}::{sheet_name}",
                rows=[row_idx],
            )
            continue
        existing = accounts[code]
        existing.amount = _sum_optional(existing.amount, amount)
        existing.rows.append(row_idx)
        existing.duplicate_count += 1
        if not existing.desc_en and desc_en:
            existing.desc_en = desc_en
        if not existing.desc_cn and desc_cn:
            existing.desc_cn = desc_cn
        if not existing.group and group:
            existing.group = group

    for account in accounts.values():
        if account.amount is not None and volume:
            account.unit = account.amount / volume * 1000
    meta = {"path": str(path), "sheet": sheet_name, "header_row": header_row, "value_col": value_col, "cpu_col": cpu_col, "volume": volume}
    wb.close()
    return accounts, meta


def extract_sap_act_eur_accounts(workbook, path: Path, sheet_name: str, month: int) -> tuple[dict[str, SourceAccount], dict[str, Any]]:
    ws = workbook[sheet_name]
    value_col = _detect_sap_month_column(ws, month)
    volume = detect_fcst_volume(workbook, month)
    summary_categories = extract_sap_summary_categories(workbook, sheet_name, month)
    accounts: dict[str, SourceAccount] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        values = list(row)
        if _is_sap_section_end(values):
            break
        code_col_idx = _find_code_index(values)
        if code_col_idx is None:
            continue
        code = account_code(values[code_col_idx])
        amount_eur = number(values[value_col - 1] if value_col - 1 < len(values) else None)
        amount = amount_eur / 1000 if amount_eur is not None else None
        desc_en = text(values[code_col_idx + 1] if code_col_idx + 1 < len(values) else "")
        unit = amount / volume * 1000 if amount is not None and volume else None
        account = SourceAccount(
            code=code,
            desc_en=desc_en,
            desc_cn="",
            group="",
            amount=amount,
            unit=unit,
            volume=volume,
            source=f"{path.name}::{sheet_name}",
            rows=[row_idx],
        )
        _add_or_sum_account(accounts, account)
    for account in accounts.values():
        if account.amount is not None and volume:
            account.unit = account.amount / volume * 1000
    meta = {"path": str(path), "sheet": sheet_name, "header_row": 4, "value_col": value_col, "cpu_col": None, "volume": volume, "amount_unit": "EUR converted to K€", "summary_categories": summary_categories}
    return accounts, meta


def extract_sap_summary_categories(workbook, sheet_name: str, month: int) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    value_col = _detect_sap_month_column(ws, month)
    buckets: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        values = list(row)
        if _is_sap_section_end(values):
            break
        raw_label = _sap_summary_key(values)
        if not raw_label:
            continue
        label = SAP_SUMMARY_LABELS.get(raw_label)
        if not label:
            continue
        amount_eur = number(values[value_col - 1] if value_col - 1 < len(values) else None)
        amount = amount_eur / 1000 if amount_eur is not None else 0.0
        if label not in buckets:
            order.append(label)
            buckets[label] = {"label": label, "amount_26": 0.0, "source_rows": [], "source_key": raw_label}
        buckets[label]["amount_26"] += amount
        buckets[label]["source_rows"].append(row_idx)
    return [buckets[label] for label in order]


def extract_4plus8_dw_categories(path: Path, month: int) -> dict[str, Any]:
    wb = safe_load_workbook(path, read_only=True, data_only=True)
    sheet_name = find_sheet(wb, ["洗碗机"])
    ws = wb[sheet_name]
    section_row = _find_row_containing(ws, "洗碗机26年实际")
    if section_row is None:
        wb.close()
        raise ValueError(f"{sheet_name} 中找不到洗碗机26年实际区块")
    header_row = _find_row_containing(ws, "科目", start=section_row, end=section_row + 5) or section_row + 1
    value_col = _find_month_value_column(ws, header_row, month)
    categories: list[dict[str, Any]] = []
    total = None
    fc_reference = None
    current_group = ""

    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 40) + 1):
        group = text(ws.cell(row_idx, 2).value)
        label = text(ws.cell(row_idx, 3).value)
        if group:
            current_group = group
        display_label = label or group
        if not display_label:
            continue
        amount = number(ws.cell(row_idx, value_col).value)
        if display_label in {"平均单价", "产量", "单台", "费率"}:
            continue
        if display_label.lower() == "functional currency":
            fc_reference = amount
            break
        if display_label == "合计":
            total = amount
            continue
        if row_idx > header_row + 15 and not group:
            continue
        categories.append(
            {
                "category": current_group,
                "label": display_label,
                "amount_26": amount,
                "source_row": row_idx,
            }
        )

    wb.close()
    return {
        "path": str(path),
        "sheet": sheet_name,
        "month": month,
        "value_col": value_col,
        "categories": categories,
        "total": total,
        "fc_reference": fc_reference,
    }


def detect_fcst_volume(workbook, month: int) -> float | None:
    sheet_name = None
    for name in workbook.sheetnames:
        if name.strip().lower() == "fcst cpu":
            sheet_name = name
            break
    if sheet_name is None:
        return None
    ws = workbook[sheet_name]
    return number(ws.cell(2, month + 1).value)


def extract_fcst_functional_currency(workbook, path: Path, month: int, volume: float | None) -> SourceAccount | None:
    sheet_name = None
    for name in workbook.sheetnames:
        if name.strip().lower() == "fcst cpu":
            sheet_name = name
            break
    if sheet_name is None:
        return None
    ws = workbook[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if text(row[0] if row else "").lower() == "functional currency impact":
            amount = number(row[month] if len(row) > month else None)
            unit = amount / volume * 1000 if amount is not None and volume else None
            return SourceAccount(
                code="__FC__",
                desc_en="Functional Currency Impact",
                desc_cn="汇率影响",
                group="Functional Currency Impact",
                amount=amount,
                unit=unit,
                volume=volume,
                source=f"{path.name}::{sheet_name}",
                rows=[row_idx],
            )
    return None


def _find_code_index(values: list[Any]) -> int | None:
    for idx, value in enumerate(values[:8]):
        if account_code(value):
            return idx
    return None


def _detect_sap_month_column(ws, month: int) -> int:
    tokens = MONTH_TOKENS[month]
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(1, min(ws.max_column, 30) + 1):
            label = text(ws.cell(row, col).value).lower()
            if any(token == label or token in label for token in tokens):
                return col
    raise ValueError(f"{ws.title} 中识别不到 {month} 月实际成本列")


def _is_functional_currency_row(values: list[Any]) -> bool:
    return any("functional currency impact" in text(value).lower() for value in values[:8])


def _is_sap_section_end(values: list[Any]) -> bool:
    return any(text(value).lower() == "over/under" for value in values[:3])


def _sap_summary_key(values: list[Any]) -> str:
    for value in values[:2]:
        label = text(value).strip()
        if not label.startswith("*"):
            continue
        return re.sub(r"^\*+\s*", "", label).strip().upper()
    return ""


def _find_row_containing(ws, needle: str, *, start: int = 1, end: int | None = None) -> int | None:
    normalized_needle = needle.lower()
    last_row = min(ws.max_row, end or ws.max_row)
    for row in range(start, last_row + 1):
        for col in range(1, min(ws.max_column, 80) + 1):
            if normalized_needle in text(ws.cell(row, col).value).lower():
                return row
    return None


def _find_month_value_column(ws, header_row: int, month: int) -> int:
    month_label = f"{month}月"
    for col in range(1, min(ws.max_column, 80) + 1):
        if month_label in text(ws.cell(header_row, col).value):
            return col
    raise ValueError(f"{ws.title} 中找不到 {month_label} 列")


def _sum_optional(left: float | None, right: float | None) -> float | None:
    if left is None and right is None:
        return None
    return (left or 0.0) + (right or 0.0)


def _add_or_sum_account(accounts: dict[str, SourceAccount], account: SourceAccount) -> None:
    if account.code not in accounts:
        accounts[account.code] = account
        return
    existing = accounts[account.code]
    existing.amount = _sum_optional(existing.amount, account.amount)
    existing.rows.extend(account.rows)
    existing.duplicate_count += account.duplicate_count
    if not existing.desc_en and account.desc_en:
        existing.desc_en = account.desc_en
    if not existing.desc_cn and account.desc_cn:
        existing.desc_cn = account.desc_cn
    if not existing.group and account.group:
        existing.group = account.group
