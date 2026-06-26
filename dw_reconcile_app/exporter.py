from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")


DETAIL_HEADERS = [
    "科目编码",
    "科目英文",
    "科目中文",
    "大科目",
    "25同期制造费",
    "25同期单台",
    "26实际制造费",
    "26实际单台",
    "单台差异",
    "额差异",
    "科目状态",
    "差异分析",
    "25来源",
    "26来源",
]


def build_export_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "科目级差异"
    _append_header(ws, DETAIL_HEADERS)
    for row in _detail_rows(payload):
        ws.append(
            [
                row.get("code"),
                row.get("desc_en"),
                row.get("desc_cn"),
                row.get("category"),
                row.get("amount_25"),
                row.get("unit_25"),
                row.get("amount_26"),
                row.get("unit_26"),
                row.get("unit_diff"),
                row.get("amount_diff"),
                row.get("status_label"),
                row.get("analysis"),
                row.get("source_25"),
                row.get("source_26"),
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    cat_ws = wb.create_sheet("大科目汇总")
    _append_header(cat_ws, ["大科目", "25同期制造费", "26实际制造费", "额差异", "25同期单台", "26实际单台", "单台差异", "科目数"])
    for item in payload.get("categories", []):
        cat_ws.append(
            [
                item.get("category"),
                item.get("amount_25"),
                item.get("amount_26"),
                item.get("amount_diff"),
                item.get("unit_25"),
                item.get("unit_26"),
                item.get("unit_diff"),
                item.get("count"),
            ]
        )
    cat_ws.freeze_panes = "A2"

    missing_ws = wb.create_sheet("未匹配科目")
    _append_header(missing_ws, DETAIL_HEADERS)
    for row in _detail_rows(payload):
        if row.get("status") in {"only_25", "only_26"}:
            missing_ws.append(
                [
                    row.get("code"),
                    row.get("desc_en"),
                    row.get("desc_cn"),
                    row.get("category"),
                    row.get("amount_25"),
                    row.get("unit_25"),
                    row.get("amount_26"),
                    row.get("unit_26"),
                    row.get("unit_diff"),
                    row.get("amount_diff"),
                    row.get("status_label"),
                    row.get("analysis"),
                    row.get("source_25"),
                    row.get("source_26"),
                ]
            )
    missing_ws.freeze_panes = "A2"

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in col) + 2))
            sheet.column_dimensions[col[0].column_letter].width = width

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _append_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def _detail_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return [row for row in payload.get("rows", []) if not row.get("is_summary")]
