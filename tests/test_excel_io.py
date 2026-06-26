from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from dw_reconcile_app.excel_io import account_code, extract_4plus8_dw_categories, extract_sap_act_eur_accounts, number


class ExcelIoTests(unittest.TestCase):
    def test_account_code_extracts_ten_digits(self):
        self.assertEqual(account_code("    6666010188  Salary - Blue collar direct employees"), "6666010188")

    def test_account_code_keeps_fix_and_var(self):
        self.assertEqual(account_code("FIX "), "FIX")
        self.assertEqual(account_code("VAR "), "VAR")

    def test_account_code_ignores_decimal_fragments(self):
        self.assertEqual(account_code(2302075.1099999994), "")
        self.assertEqual(account_code("2412982.9200000004"), "")

    def test_account_code_accepts_numeric_ten_digit_codes(self):
        self.assertEqual(account_code(6666110101), "6666110101")
        self.assertEqual(account_code(6666110101.0), "6666110101")

    def test_account_code_accepts_six_digit_cost_elements(self):
        self.assertEqual(account_code(600006), "600006")
        self.assertEqual(account_code("    600006      OVH Administrative"), "600006")

    def test_number_handles_blank_and_numeric_text(self):
        self.assertIsNone(number(""))
        self.assertIsNone(number("#DIV/0!"))
        self.assertEqual(number("1,234.5"), 1234.5)

    def test_sap_act_eur_stops_at_first_over_under_section(self):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "SAP_ACT_EUR"
        ws.cell(2, 3).value = "Jan"
        ws.cell(5, 1).value = "6666010188"
        ws.cell(5, 2).value = "Salary - Blue collar direct employees"
        ws.cell(5, 3).value = 1000
        ws.cell(6, 1).value = "*   CS_Depreciation"
        ws.cell(6, 2).value = "*   CS_Depreciation"
        ws.cell(6, 3).value = 2000
        ws.cell(7, 1).value = "Over/Under"
        ws.cell(8, 1).value = "6666010188"
        ws.cell(8, 2).value = "Duplicate section that must be ignored"
        ws.cell(8, 3).value = 5000

        fcst = workbook.create_sheet("FCST CPU ")
        fcst.cell(2, 2).value = 100

        accounts, metadata = extract_sap_act_eur_accounts(workbook, Path("sample.xlsx"), "SAP_ACT_EUR", 1)

        self.assertAlmostEqual(accounts["6666010188"].amount, 1.0)
        self.assertEqual(accounts["6666010188"].rows, [5])
        self.assertEqual(metadata["summary_categories"][0]["label"], "折旧（含FC）")
        self.assertEqual(metadata["summary_categories"][0]["amount_26"], 2.0)

    def test_sap_act_eur_does_not_add_fcst_functional_currency_as_account(self):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "SAP_ACT_EUR"
        ws.cell(2, 3).value = "Jan"
        ws.cell(5, 1).value = "6666110101"
        ws.cell(5, 2).value = "Depreciation of buildings"
        ws.cell(5, 3).value = 1000
        ws.cell(6, 1).value = "Over/Under"
        fcst = workbook.create_sheet("FCST CPU ")
        fcst.cell(2, 2).value = 100
        fcst.cell(21, 1).value = "Functional Currency Impact"
        fcst.cell(21, 2).value = 5

        accounts, _metadata = extract_sap_act_eur_accounts(workbook, Path("sample.xlsx"), "SAP_ACT_EUR", 1)

        self.assertNotIn("__FC__", accounts)

    def test_extract_4plus8_dw_categories_reads_26_actual_section(self):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "洗碗机"
        ws.cell(76, 3).value = "洗碗机26年实际"
        ws.cell(77, 3).value = "科目"
        ws.cell(77, 4).value = "25年1月"
        ws.cell(80, 2).value = "固定制造"
        ws.cell(80, 3).value = "折旧（含FC）"
        ws.cell(80, 4).value = 564.25121
        ws.cell(96, 2).value = "合计"
        ws.cell(96, 4).value = 1778.33998
        ws.cell(98, 2).value = "functional currency"
        ws.cell(98, 4).value = 392.90298
        ws.cell(101, 3).value = "洗碗机26年预算"
        ws.cell(102, 3).value = "科目"
        ws.cell(102, 4).value = "25年1月"
        ws.cell(105, 2).value = "固定制造"
        ws.cell(105, 3).value = "折旧（含FC）"
        ws.cell(105, 4).value = 999

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "ref.xlsx"
            workbook.save(path)
            result = extract_4plus8_dw_categories(path, month=1)

        self.assertEqual(result["total"], 1778.33998)
        self.assertEqual(result["fc_reference"], 392.90298)
        self.assertEqual(result["categories"][0]["category"], "固定制造")
        self.assertEqual(result["categories"][0]["label"], "折旧（含FC）")
        self.assertEqual(result["categories"][0]["amount_26"], 564.25121)
        self.assertNotIn(999, [item["amount_26"] for item in result["categories"]])


if __name__ == "__main__":
    unittest.main()
