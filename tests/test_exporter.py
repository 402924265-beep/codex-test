from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from dw_reconcile_app.exporter import build_export_workbook


class ExporterTests(unittest.TestCase):
    def test_export_workbook_contains_detail_and_category_sheets(self):
        payload = {
            "rows": [
                {
                    "code": "6666010188",
                    "desc_en": "Salary",
                    "desc_cn": "工资",
                    "category": "直接人工",
                    "amount_25": 1,
                    "unit_25": 2,
                    "amount_26": 3,
                    "unit_26": 4,
                    "unit_diff": 2,
                    "amount_diff": 2,
                    "status": "both",
                    "status_label": "两边都有",
                    "analysis": "测试分析",
                    "source_25": "a",
                    "source_26": "b",
                }
            ],
            "categories": [{"category": "直接人工", "amount_25": 1, "amount_26": 3, "amount_diff": 2, "unit_25": 2, "unit_26": 4, "unit_diff": 2, "count": 1}],
        }
        content = build_export_workbook(payload)
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        self.assertIn("科目级差异", wb.sheetnames)
        self.assertIn("大科目汇总", wb.sheetnames)
        self.assertEqual(wb["科目级差异"]["L2"].value, "测试分析")
        wb.close()

    def test_export_workbook_excludes_fix_var_summary_rows_from_detail(self):
        payload = {
            "rows": [
                {"code": "FIX", "desc_en": "* CS_Depreciation", "is_summary": True, "status": "only_25"},
                {"code": "6666110101", "desc_en": "Depreciation", "is_summary": False, "status": "both"},
            ],
            "categories": [],
        }
        content = build_export_workbook(payload)
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb["科目级差异"]
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["A2"].value, "6666110101")
        wb.close()


if __name__ == "__main__":
    unittest.main()
